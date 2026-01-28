import streamlit as st
import pandas as pd
from io import BytesIO
from datetime import datetime
from openpyxl.styles import Font, PatternFill, Alignment
from PIL import Image

# --- CONFIGURACIÓN DE LA PÁGINA ---
# Usamos un icono más corporativo y un título más formal en la pestaña del navegador
st.set_page_config(
    page_title="Portal de Inventarios", 
    page_icon="📈", 
    layout="centered",
    initial_sidebar_state="collapsed"
)

# --- ESTILOS CSS PERSONALIZADOS (Para un toque extra de limpieza) ---
# Esto oculta el menú de hamburguesa de Streamlit y el pie de página para que se vea más como una app propia.
hide_streamlit_style = """
            <style>
            #MainMenu {visibility: hidden;}
            footer {visibility: hidden;}
            header {visibility: hidden;}
            </style>
            """
st.markdown(hide_streamlit_style, unsafe_allow_html=True)

# --- COLORES EXCEL ---
COLOR_ENCABEZADO = "B4C6E7"

# --- FUNCIONES AUXILIARES ---
def obtener_fecha_leible():
    meses = ("ENERO", "FEBRERO", "MARZO", "ABRIL", "MAYO", "JUNIO", "JULIO", "AGOSTO", "SEPTIEMBRE", "OCTUBRE", "NOVIEMBRE", "DICIEMBRE")
    hoy = datetime.now()
    return f"{hoy.day} {meses[hoy.month-1]} {hoy.year}"

# ==========================================
# INTERFAZ DE USUARIO PROFESIONAL
# ==========================================

# 1. ENCABEZADO CON LOGO (Usamos columnas para organizarlo)
# Crea dos columnas: una estrecha para el logo, una ancha para el título
col_logo, col_titulo = st.columns([1, 4])

with col_logo:
    # INTENTA CARGAR EL LOGO. Si no existe, no pasa nada.
    try:
        image = Image.open('logo.png') # Asegúrate que tu imagen se llame así
        st.image(image, width=120)
    except FileNotFoundError:
        st.warning("⚠️ Falta el archivo 'logo.png' en la carpeta.")

with col_titulo:
    st.title("Sistema de Gestión de Inventarios")
    st.markdown("##### Portal de procesamiento y cálculo de precios.")

st.markdown("---") # Línea separadora

# 2. ÁREA DE INSTRUCCIONES (En un recuadro azul limpio)
with st.container():
    st.info(
        """
        ℹ️ **Instrucciones de uso:**
        
        1. Descargue el archivo maestro desde su sistema ERP.
        2. Arrastre el archivo **'Valor del inventario por clasificación.xlsx'** en el recuadro inferior.
        3. El sistema procesará automáticamente las sucursales y calculará los precios de lista y promoción.
        """
    )

st.write("") # Espacio en blanco

# 3. WIDGET DE CARGA DE ARCHIVO (Más prominente)
st.markdown("### 📂 Cargar Archivo Maestro")
archivo_subido = st.file_uploader("", type=["xlsx"], help="Solo se permiten archivos Excel (.xlsx)")

# ==========================================
# LÓGICA DE PROCESAMIENTO
# ==========================================
if archivo_subido is not None:
    # Usamos un 'spinner' para mostrar que está trabajando
    with st.spinner('🔄 Procesando datos, separando almacenes y aplicando formato... Por favor espere.'):
        try:
            # --- PROCESO DE DATOS (En memoria) ---
            df_completo = pd.read_excel(archivo_subido, header=0)
            columna_codigo = df_completo['Código'].astype(str).str.upper()
            
            # Búsqueda de "LLANTAS"
            filas_llantas = df_completo.index[columna_codigo.str.contains('LLANTAS', na=False)].tolist()

            if len(filas_llantas) < 2:
                st.error("❌ Error crítico: La estructura del Excel no es válida. No se encontró el separador 'LLANTAS' las veces necesarias.")
                st.stop()
            
            fila_corte = filas_llantas[1]
            df_aburto_raw = df_completo.iloc[:fila_corte].copy()
            df_mayoreo_raw = df_completo.iloc[fila_corte+1:].copy()

            def procesar_almacen(df):
                df['Existencia'] = pd.to_numeric(df['Existencia'], errors='coerce')
                df_limpio = df.dropna(subset=['Existencia']).copy()
                df_limpio['Costo unitario'] = pd.to_numeric(df_limpio['Costo unitario'], errors='coerce').fillna(0)
                
                df_final = pd.DataFrame()
                df_final['Código'] = df_limpio['Código']
                df_final['Descripcion'] = df_limpio['Artículo']
                df_final['Existencia'] = df_limpio['Existencia']
                df_final['Precio lista'] = (df_limpio['Costo unitario'] * 1.40)
                df_final['Promocion'] = (df_limpio['Costo unitario'] * 1.25)
                df_final['Remate'] = None
                return df_final

            df_mayoreo = procesar_almacen(df_mayoreo_raw)
            df_aburto = procesar_almacen(df_aburto_raw)

            # --- GUARDAR EN MEMORIA (Buffer) ---
            buffer = BytesIO()
            writer = pd.ExcelWriter(buffer, engine='openpyxl')
            df_mayoreo.to_excel(writer, sheet_name='MAYOREO HERMOSILLO', startrow=5, index=False)
            df_aburto.to_excel(writer, sheet_name='SUCURSAL ABURTO', startrow=5, index=False)

            # --- MAQUILLAJE EXCEL ---
            def maquillar(ws, nombre_sucursal):
                titulos = [f"INVENTARIO {obtener_fecha_leible()}", nombre_sucursal, "PRECIOS IVA INCLUIDO", "CONTACTO..."]
                for i, t in enumerate(titulos, 1):
                    ws.cell(row=i, column=1, value=t).font = Font(bold=True, size=11)
                
                fill = PatternFill(start_color=COLOR_ENCABEZADO, fill_type="solid")
                for col in range(1, 7):
                    cell = ws.cell(row=6, column=col)
                    cell.fill = fill
                    cell.font = Font(bold=True)
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                
                for row in range(7, ws.max_row + 1):
                    ws.cell(row=row, column=1).alignment = Alignment(horizontal='left')
                    ws.cell(row=row, column=3).alignment = Alignment(horizontal='center')
                    for c in [4, 5, 6]:
                        ws.cell(row=row, column=c).number_format = '"$"#,##0.00'

                ws.column_dimensions['A'].width = 25  
                ws.column_dimensions['B'].width = 65  
                ws.column_dimensions['C'].width = 15  
                ws.column_dimensions['D'].width = 20  
                ws.column_dimensions['E'].width = 20  
                ws.column_dimensions['F'].width = 20  

            maquillar(writer.sheets['MAYOREO HERMOSILLO'], "CEDIS HERMOSILLO")
            maquillar(writer.sheets['SUCURSAL ABURTO'], "ALMACEN ABURTO")
            writer.close()
            buffer.seek(0)

        except Exception as e:
            st.error(f"Ocurrió un error inesperado: {e}")
            st.stop()

    # 4. ÁREA DE DESCARGA (Se muestra al finalizar con éxito)
    st.markdown("---")
    st.success("✅ ¡Archivo procesado exitosamente!")
    
    col_izq, col_centro, col_der = st.columns([1,2,1])
    with col_centro:
        fecha_hoy = datetime.now().strftime("%Y-%m-%d")
        st.download_button(
            label="⬇️ DESCARGAR INVENTARIO FINAL",
            data=buffer,
            file_name=f"Inventario_Procesado_{fecha_hoy}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True # Hace el botón ancho y llamativo
        )